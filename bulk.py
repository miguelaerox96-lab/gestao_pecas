import os
import shutil
import zipfile
import json
import uuid
import openpyxl
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from tempfile import mkdtemp
from typing import List, Dict, Optional, Any
import uuid

import models
from database import get_db
from routers.auth import check_admin, get_download_admin

router = APIRouter(tags=["bulk"])

def remove_file(path: str):
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception as e:
        print(f"Erro ao remover ficheiro temporário {path}: {e}")

# 1. TEMPLATE DOWNLOAD
@router.get("/bulk/template/{type_id}")
def download_template(type_id: int, background_tasks: BackgroundTasks, db: Session = Depends(get_db), admin: models.User = Depends(get_download_admin)):
    part_type = db.query(models.PartType).filter(models.PartType.id == type_id).first()
    if not part_type:
        raise HTTPException(status_code=404, detail="Type not found")
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Template {part_type.name}"
    
    headers = [
        "REF_PECA", "LOCALIZACAO", "MARCA", "MODELO", "ANO", 
        "PRECO", "MOSTRAR_PRECO_SITE", "OBSERVACOES", "FOTOS_FILENAMES"
    ]
    
    dynamic_fields = [f.name for f in part_type.fields]
    headers.extend(dynamic_fields)
    
    ws.append(["--- INSTRUÇÕES ---"])
    ws.append(["1. Preencha os dados abaixo respeitando os cabeçalhos."])
    ws.append(["2. Em FOTOS_FILENAMES coloque o nome exato dos ficheiros separados por vírgula (Ex: img1.jpg, foto.png)."])
    ws.append(["3. Junte este Excel com uma pasta chamada 'imagens' contendo as fotos num ficheiro ZIP e faça upload."])
    ws.append([f"TYPE_ID={part_type.id} (NÃO APAGAR ESTA LINHA)"])
    ws.append([]) # Empty line
    
    ws.append(headers)
    
    # Format Headers (Bold)
    header_row_index = 6
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=header_row_index, column=col_num).font = openpyxl.styles.Font(bold=True)
    
    os.makedirs("storage/temp", exist_ok=True)
    temp_file = f"storage/temp/template_{uuid.uuid4()}.xlsx"
    wb.save(temp_file)
    
    background_tasks.add_task(remove_file, temp_file)
    
    return FileResponse(
        path=temp_file, 
        filename=f"Template_{part_type.name}.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# 2. ZIP IMPORT
@router.post("/bulk/import")
def import_parts_zip(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.User = Depends(check_admin)):
    if not file.filename.endswith(".zip"):
        raise HTTPException(status_code=400, detail="Must be a ZIP file")
        
    work_dir = mkdtemp()
    zip_path = os.path.join(work_dir, "upload.zip")
    
    with open(zip_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(work_dir)
        
    # Analyze the contents
    extracted_items = os.listdir(work_dir)
    excel_file = next((f for f in extracted_items if f.endswith('.xlsx')), None)
    
    if not excel_file:
        shutil.rmtree(work_dir)
        raise HTTPException(status_code=400, detail="O ficheiro ZIP não contem nenhum ficheiro Excel (.xlsx)")
        
    img_dir_candidates = [os.path.join(work_dir, d) for d in extracted_items if os.path.isdir(os.path.join(work_dir, d)) and d.lower() == "imagens"]
    img_dir = img_dir_candidates[0] if img_dir_candidates else None
        
    wb = openpyxl.load_workbook(os.path.join(work_dir, excel_file))
    ws = wb.active
    
    results = {"success": 0, "errors": []}
    
    # Identify type_id
    type_id = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] and str(row[0]).startswith("TYPE_ID="):
            try:
                type_id = int(str(row[0]).split("=")[1].strip())
            except:
                pass
                
    if not type_id:
        shutil.rmtree(work_dir)
        raise HTTPException(status_code=400, detail="Falta a tag TYPE_ID= no topo do template.")
            
    # Find headers row
    header_idx = -1
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == "REF_PECA":
            header_idx = i
            break
            
    if header_idx == -1:
        shutil.rmtree(work_dir)
        raise HTTPException(status_code=400, detail="Não conseguimos encontrar a linha de cabeçalho (que começa em REF_PECA)")
        
    headers = [str(c).strip() if c else "" for c in list(ws.iter_rows(values_only=True))[header_idx]]
    
    # Process Rows
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_idx+2, values_only=True), start=header_idx+2):
        if not row[0]: continue # Empty row
        
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        
        ref = row_dict.get("REF_PECA", "")
        loc = row_dict.get("LOCALIZACAO", "")
        
        if not ref or not loc:
            results["errors"].append({"row": r_idx, "msg": "REF_PECA e LOCALIZACAO são obrigatórios."})
            continue

        # Check existing Part in this location
        existing = db.query(models.Part).filter(
            models.Part.part_number == str(ref), 
            models.Part.type_id == type_id
        ).first()

        if existing and existing.status != "EmptySlot":
             results["errors"].append({"row": r_idx, "msg": f"A Ref {ref} já existe para este tipo e está Disponível."})
             continue

        # Extract dynamic data
        dyn_data = {}
        for k, v in row_dict.items():
            if k not in ["REF_PECA", "LOCALIZACAO", "MARCA", "MODELO", "ANO", "PRECO", "MOSTRAR_PRECO_SITE", "OBSERVACOES", "FOTOS_FILENAMES"]:
                if v is not None:
                    dyn_data[k] = str(v)
                    
        # Manage Images
        images_urls = []
        fotos_str = row_dict.get("FOTOS_FILENAMES", "")
        if fotos_str and img_dir:
            file_names = [x.strip() for x in str(fotos_str).split(",")]
            for fn in file_names:
                p_file = os.path.join(img_dir, fn)
                if os.path.exists(p_file):
                    unique_fn = f"{int(datetime.now().timestamp())}_{fn}"
                    dest = os.path.join("storage", unique_fn)
                    if not os.path.exists("storage"): os.makedirs("storage")
                    shutil.copy2(p_file, dest)
                    images_urls.append(f"/storage/{unique_fn}")
        
        # Determine Show Price boolean
        show_str = str(row_dict.get("MOSTRAR_PRECO_SITE", "")).strip().lower()
        show_price = show_str not in ["nao", "não", "false", "0"]
        
        # Create or Update Part
        try:
            if existing:
                # Recycle EmptySlot
                existing.location = str(loc)
                existing.brand = str(row_dict.get("MARCA", "")) if row_dict.get("MARCA") else None
                existing.model = str(row_dict.get("MODELO", "")) if row_dict.get("MODELO") else None
                existing.year = str(row_dict.get("ANO", "")) if row_dict.get("ANO") else None
                existing.price = str(row_dict.get("PRECO", "")) if row_dict.get("PRECO") else None
                existing.show_price = show_price
                existing.description = str(row_dict.get("OBSERVACOES", "")) if row_dict.get("OBSERVACOES") else None
                existing.images = images_urls
                existing.status = "Available"
                existing.dynamic_data = dyn_data
                
                db.flush()
                hist = models.HistoryRecord(part_id=existing.id, action="Restocked (Bulk)", price_at_action=existing.price, user=admin.username)
                db.add(hist)
            else:
                new_part = models.Part(
                    part_number=str(ref),
                    location=str(loc),
                    type_id=type_id,
                    brand=str(row_dict.get("MARCA", "")) if row_dict.get("MARCA") else None,
                    model=str(row_dict.get("MODELO", "")) if row_dict.get("MODELO") else None,
                    year=str(row_dict.get("ANO", "")) if row_dict.get("ANO") else None,
                    price=str(row_dict.get("PRECO", "")) if row_dict.get("PRECO") else None,
                    show_price=show_price,
                    description=str(row_dict.get("OBSERVACOES", "")) if row_dict.get("OBSERVACOES") else None,
                    images=images_urls,
                    status="Available",
                    dynamic_data=dyn_data
                )
                db.add(new_part)
                db.flush()
                hist = models.HistoryRecord(part_id=new_part.id, action="Created (Bulk)", price_at_action=new_part.price, user=admin.username)
                db.add(hist)
            results["success"] += 1
        except Exception as e:
            results["errors"].append({"row": r_idx, "msg": f"Erro interno: {str(e)}"})

    db.commit()
    shutil.rmtree(work_dir)
    return results


# 3. EXPORT INVENTORY
@router.get("/bulk/export/inventory")
def export_inventory(background_tasks: BackgroundTasks, db: Session = Depends(get_db), admin: models.User = Depends(get_download_admin)):
    parts = db.query(models.Part).filter(models.Part.status == "Available").all()
    types = {t.id: t.name for t in db.query(models.PartType).all()}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario AutoParts"
    
    headers = ["ID", "TIPO", "REFERENCIA", "MARCA", "MODELO", "ANO", "PRECO", "PRECO_OCULTO", "LOCALIZACAO"]
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
        
    for p in parts:
        ws.append([
            p.id,
            types.get(p.type_id, "Desconhecido"),
            p.part_number,
            p.brand or "-",
            p.model or "-",
            p.year or "-",
            p.price or "0.00",
            "Sim" if not p.show_price else "Nao",
            p.location
        ])
        
    os.makedirs("storage/temp", exist_ok=True)
    out_path = f"storage/temp/export_inventory_{uuid.uuid4()}.xlsx"
    wb.save(out_path)
    
    background_tasks.add_task(remove_file, out_path)
    
    return FileResponse(
        path=out_path, 
        filename=f"Inventario_AutoParts_{datetime.now().strftime('%Y%m%d')}.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# 4. EXPORT HISTORY (SALES)
@router.get("/bulk/export/history")
def export_history(background_tasks: BackgroundTasks, db: Session = Depends(get_db), admin: models.User = Depends(get_download_admin)):
    records = db.query(models.HistoryRecord).order_by(models.HistoryRecord.timestamp.desc()).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registo de Historico"
    
    headers = ["DATA_HORA", "ACAO", "ID_PECA", "PRECO_NO_MOMENTO", "DETALHES"]
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
        
    total_sales = 0
    revenue = 0.0
        
    for r in records:
        ws.append([
            str(r.timestamp),
            r.action,
            r.part_id,
            r.price_at_action or "-",
            r.details or ""
        ])
        
        if r.action == "Sold":
            total_sales += 1
            if r.price_at_action:
                try:
                    revenue += float(r.price_at_action)
                except:
                    pass
                    
    # Blank row
    ws.append([])
    ws.append(["RESUMO FINANCEIRO"])
    ws.cell(row=ws.max_row, column=1).font = openpyxl.styles.Font(bold=True)
    ws.append(["Total Vendido:", total_sales])
    ws.append(["Receita Total Calculada (€):", revenue])
    
    os.makedirs("storage/temp", exist_ok=True)
    out_path = f"storage/temp/export_history_{uuid.uuid4()}.xlsx"
    wb.save(out_path)
    
    background_tasks.add_task(remove_file, out_path)
    
    return FileResponse(
        path=out_path, 
        filename=f"Historico_AutoParts_{datetime.now().strftime('%Y%m%d')}.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 5. BACKUP EXPORT
@router.get("/backup/export")
def export_backup(format: Optional[str] = "both", background_tasks: BackgroundTasks = None, db: Session = Depends(get_db), admin: models.User = Depends(get_download_admin)):
    # Tables to export
    tables = {
        "brands": [{"id": x.id, "name": x.name} for x in db.query(models.Brand).all()],
        "locations": [{"id": x.id, "name": x.name} for x in db.query(models.Location).all()],
        "part_types": [{"id": x.id, "name": x.name} for x in db.query(models.PartType).all()],
        "type_fields": [{"id": x.id, "part_type_id": x.part_type_id, "name": x.name, "field_type": x.field_type, "options": x.options, "keep_on_baixa": x.keep_on_baixa, "required_field": x.required_field} for x in db.query(models.TypeField).all()],
        "parts": [{"id": x.id, "part_number": x.part_number, "location": x.location, "type_id": x.type_id, "brand": x.brand, "model": x.model, "year": x.year, "price": x.price, "show_price": x.show_price, "description": x.description, "images": x.images, "status": x.status, "dynamic_data": x.dynamic_data} for x in db.query(models.Part).all()],
        "vehicles": [{"id": x.id, "vin": x.vin, "make": x.make, "model": x.model, "year": x.year, "vehicle_type": x.vehicle_type, "price": x.price, "show_price": x.show_price, "description": x.description, "engine": x.engine, "mileage": x.mileage, "images": x.images, "status": x.status} for x in db.query(models.Vehicle).all()],
        "inquiries": [{"id": x.id, "part_id": x.part_id, "vehicle_id": x.vehicle_id, "email": x.email, "phone": x.phone, "message": x.message, "created_at": x.created_at.isoformat(), "status": x.status} for x in db.query(models.Inquiry).all()],
        "history": [{"id": x.id, "part_id": x.part_id, "vehicle_id": x.vehicle_id, "action": x.action, "timestamp": x.timestamp.isoformat(), "price_at_action": x.price_at_action, "details": x.details} for x in db.query(models.HistoryRecord).all()]
    }

    work_dir = mkdtemp()
    
    # 3. Zipping
    zip_path = os.path.join(work_dir, "Backup_AutoParts.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        
        # Add Excel if requested
        if format in ["excel", "both"]:
            wb = openpyxl.Workbook()
            first = True
            for name, data in tables.items():
                if first:
                    ws = wb.active
                    ws.title = name
                    first = False
                else:
                    ws = wb.create_sheet(title=name)
                
                if data:
                    headers = list(data[0].keys())
                    ws.append(headers)
                    for row in data:
                        values = []
                        for h in headers:
                            val = row[h]
                            if isinstance(val, (list, dict)):
                                val = json.dumps(val, ensure_ascii=False)
                            values.append(val)
                        ws.append(values)
            excel_path = os.path.join(work_dir, "data.xlsx")
            wb.save(excel_path)
            zipf.write(excel_path, "data.xlsx")

        # Add JSON if requested
        if format in ["json", "both"]:
            json_path = os.path.join(work_dir, "data.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(tables, f, ensure_ascii=False, indent=2)
            zipf.write(json_path, "data.json")

        # Always include storage folder
        if os.path.exists("storage"):
            for root, dirs, files in os.walk("storage"):
                for file in files:
                    if file.endswith((".jpg", ".png", ".webp", ".jpeg", ".ico", ".svg")):
                        file_path = os.path.join(root, file)
                        if not file.startswith(("template_", "export_", "Backup_")):
                            zipf.write(file_path, os.path.join("storage", file))
                        
    os.makedirs("storage/temp", exist_ok=True)
    final_zip = f"storage/temp/Backup_AutoParts_{uuid.uuid4()}.zip"
    shutil.copy2(zip_path, final_zip)
    shutil.rmtree(work_dir)
    
    background_tasks.add_task(remove_file, final_zip)
    
    return FileResponse(path=final_zip, filename=f"AutoParts_Backup_{format}_{datetime.now().strftime('%Y%m%d_%H%M')}.zip", media_type="application/zip")

@router.get("/backup/export/json")
def export_backup_json(background_tasks: BackgroundTasks, db: Session = Depends(get_db), admin: models.User = Depends(get_download_admin)):
    # Tables to export
    tables = {
        "brands": [{"id": x.id, "name": x.name} for x in db.query(models.Brand).all()],
        "locations": [{"id": x.id, "name": x.name} for x in db.query(models.Location).all()],
        "part_types": [{"id": x.id, "name": x.name} for x in db.query(models.PartType).all()],
        "type_fields": [{"id": x.id, "part_type_id": x.part_type_id, "name": x.name, "field_type": x.field_type, "options": x.options, "keep_on_baixa": x.keep_on_baixa, "required_field": x.required_field} for x in db.query(models.TypeField).all()],
        "parts": [{"id": x.id, "part_number": x.part_number, "location": x.location, "type_id": x.type_id, "brand": x.brand, "model": x.model, "year": x.year, "price": x.price, "show_price": x.show_price, "description": x.description, "images": x.images, "status": x.status, "dynamic_data": x.dynamic_data} for x in db.query(models.Part).all()],
        "vehicles": [{"id": x.id, "vin": x.vin, "make": x.make, "model": x.model, "year": x.year, "vehicle_type": x.vehicle_type, "price": x.price, "show_price": x.show_price, "description": x.description, "engine": x.engine, "mileage": x.mileage, "images": x.images, "status": x.status} for x in db.query(models.Vehicle).all()],
        "inquiries": [{"id": x.id, "part_id": x.part_id, "vehicle_id": x.vehicle_id, "email": x.email, "phone": x.phone, "message": x.message, "created_at": x.created_at.isoformat(), "status": x.status} for x in db.query(models.Inquiry).all()],
        "history": [{"id": x.id, "part_id": x.part_id, "vehicle_id": x.vehicle_id, "action": x.action, "timestamp": x.timestamp.isoformat(), "price_at_action": x.price_at_action, "details": x.details} for x in db.query(models.HistoryRecord).all()]
    }
    
    os.makedirs("storage/temp", exist_ok=True)
    file_path = f"storage/temp/data_backup_{uuid.uuid4()}.json"
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(tables, f, ensure_ascii=False, indent=2)
        
    background_tasks.add_task(remove_file, file_path)
        
    return FileResponse(path=file_path, filename=f"AutoParts_Data_{datetime.now().strftime('%Y%m%d_%H%M')}.json", media_type="application/json")

# 6. BACKUP IMPORT
@router.post("/backup/import")
def import_backup(mode: str = "replace", file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.User = Depends(check_admin)):
    work_dir = mkdtemp()
    data = {}
    
    try:
        # 1. Load Data
        if file.filename.endswith(".json"):
            data = json.load(file.file)
        elif file.filename.endswith(".zip"):
            zip_path = os.path.join(work_dir, "upload.zip")
            with open(zip_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(work_dir)
                
            json_path = os.path.join(work_dir, "data.json")
            excel_path = os.path.join(work_dir, "data.xlsx")
            
            if os.path.exists(json_path):
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            elif os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows: 
                        data[sheet_name] = []
                        continue
                    headers = rows[0]
                    sheet_data = []
                    for row in rows[1:]:
                        d = {}
                        for i, h in enumerate(headers):
                            val = row[i]
                            if isinstance(val, str) and (val.startswith('[') or val.startswith('{')):
                                try: val = json.loads(val)
                                except: pass
                            d[h] = val
                        sheet_data.append(d)
                    data[sheet_name] = sheet_data
            
            # Restore storage images if from ZIP
            storage_extract_dir = os.path.join(work_dir, "storage")
            if os.path.exists(storage_extract_dir):
                if not os.path.exists("storage"): os.makedirs("storage")
                for file_name in os.listdir(storage_extract_dir):
                    src = os.path.join(storage_extract_dir, file_name)
                    dest = os.path.join("storage", file_name)
                    shutil.copy2(src, dest)
        else:
            raise HTTPException(status_code=400, detail="Formato não suportado. Use .zip ou .json.")

        if not data:
            raise HTTPException(status_code=400, detail="Nenhum dado encontrado no ficheiro.")

        # 2. Process Data
        if mode == "replace":
            # Extremely destructive!
            db.query(models.HistoryRecord).delete()
            db.query(models.Inquiry).delete()
            db.query(models.Part).delete()
            db.query(models.Vehicle).delete()
            db.query(models.TypeField).delete()
            db.query(models.PartType).delete()
            db.query(models.Location).delete()
            db.query(models.Brand).delete()
            db.commit()
            
            for j in data.get("brands", []): db.add(models.Brand(**j))
            for j in data.get("locations", []): db.add(models.Location(**j))
            for j in data.get("part_types", []): 
                db.add(models.PartType(**{k:v for k,v in j.items() if k != 'fields'}))
            db.commit() 
            
            for j in data.get("type_fields", []):
                db.add(models.TypeField(
                    id=j.get("id"),
                    part_type_id=j.get("part_type_id"),
                    name=j.get("name"),
                    field_type=j.get("field_type", "text"),
                    options=j.get("options", []),
                    keep_on_baixa=j.get("keep_on_baixa", False),
                    required_field=j.get("required_field", False)  # safe default for old backups
                ))
            for j in data.get("parts", []): db.add(models.Part(**j))
            for j in data.get("vehicles", []): db.add(models.Vehicle(**j))
            
            for j in data.get("inquiries", []):
                if 'created_at' in j: j['created_at'] = datetime.fromisoformat(j['created_at'])
                db.add(models.Inquiry(**j))
            
            for j in data.get("history", []):
                if 'timestamp' in j: j['timestamp'] = datetime.fromisoformat(j['timestamp'])
                db.add(models.HistoryRecord(**j))
            db.commit()
            
        else: # merge
            for b in data.get("brands", []):
                if not db.query(models.Brand).filter_by(name=b["name"]).first(): db.add(models.Brand(name=b["name"]))
            for l in data.get("locations", []):
                if not db.query(models.Location).filter_by(name=l["name"]).first(): db.add(models.Location(name=l["name"]))
            db.commit()

            # Merge part types and fields
            type_id_map = {}  # old_id -> new_id
            for pt in data.get("part_types", []):
                existing_pt = db.query(models.PartType).filter_by(name=pt["name"]).first()
                if not existing_pt:
                    new_pt = models.PartType(name=pt["name"])
                    db.add(new_pt)
                    db.flush()
                    type_id_map[pt["id"]] = new_pt.id
                else:
                    type_id_map[pt["id"]] = existing_pt.id
            db.commit()

            for tf in data.get("type_fields", []):
                mapped_type_id = type_id_map.get(tf["part_type_id"], tf["part_type_id"])
                existing_tf = db.query(models.TypeField).filter_by(part_type_id=mapped_type_id, name=tf["name"]).first()
                if not existing_tf:
                    db.add(models.TypeField(
                        part_type_id=mapped_type_id,
                        name=tf["name"],
                        field_type=tf.get("field_type", "text"),
                        options=tf.get("options", []),
                        keep_on_baixa=tf.get("keep_on_baixa", False),
                        required_field=tf.get("required_field", False)
                    ))
            db.commit()
            
            for p in data.get("parts", []):
                if not db.query(models.Part).filter_by(part_number=p["part_number"], type_id=p["type_id"]).first():
                    db.add(models.Part(**{k:v for k,v in p.items() if k != "id"}))
            
            for v in data.get("vehicles", []):
                if not db.query(models.Vehicle).filter_by(vin=v["vin"]).first():
                    db.add(models.Vehicle(**{k:v for k,v in v.items() if k != "id"}))
            db.commit()

        shutil.rmtree(work_dir)
        return {"msg": f"Restaurado com Sucesso (Modo: {mode})"}
    except Exception as e:
        if os.path.exists(work_dir): shutil.rmtree(work_dir)
        raise HTTPException(status_code=500, detail=f"Restore failed: {str(e)}")


@router.post("/bulk/parts")
def bulk_create_parts(parts_in: List[Dict[str, Any]], db: Session = Depends(get_db), admin: models.User = Depends(check_admin)):
    """
    Novo endpoint para inserção rápida em lote. 
    Recebe uma lista de dicionários mapeados para o modelo Part.
    """
    results = {"success": 0, "skipped": 0, "errors": []}
    
    # Obter tipos de peças para validar/cache
    type_ids = {t.id for t in db.query(models.PartType.id).all()}
    
    for idx, p_data in enumerate(parts_in):
        try:
            type_id = p_data.get("type_id")
            pn = str(p_data.get("part_number", "")).strip()
            
            if not pn or type_id not in type_ids:
                results["errors"].append({"idx": idx, "msg": "Dados inválidos: part_number ou type_id em falta/inválido."})
                continue
                
            # Verificar se já existe (para evitar erro de UniqueConstraint)
            existing = db.query(models.Part).filter(
                models.Part.part_number == pn,
                models.Part.type_id == type_id
            ).first()
            
            if existing:
                if existing.status == "EmptySlot":
                    # Reutilizar slot
                    for k, v in p_data.items():
                        if k != "id": setattr(existing, k, v)
                    existing.status = "Available"
                    results["success"] += 1
                else:
                    results["skipped"] += 1
                continue

            # Criar nova peça
            new_part = models.Part(
                part_number=pn,
                type_id=type_id,
                brand=p_data.get("brand"),
                model=p_data.get("model"),
                year=p_data.get("year"),
                price=p_data.get("price"),
                location=p_data.get("location", ""),
                description=p_data.get("description"),
                dynamic_data=p_data.get("dynamic_data", {}),
                images=p_data.get("images", []),
                show_price=p_data.get("show_price", True),
                status="Available"
            )
            # History for bulk creation
            hist = models.HistoryRecord(part_id=new_part.id, action="Created (Bulk API)", price_at_action=new_part.price, user=admin.username)
            db.add(hist)

            # Commit a cada batch de 100 para não estourar memória se a lista for gigante
            if results["success"] % 100 == 0:
                db.flush()
                
        except Exception as e:
            results["errors"].append({"idx": idx, "msg": str(e)})

    db.commit()
    return results
